"""Excel出力モジュール（SDU_shift テンプレート流し込み方式）

export/templates/SDU_shift_template.xlsx を読み込み、生成済みシフトデータを
値として流し込む。書式・マージ・列幅・印刷設定はテンプレートのものを保持する。

テンプレートのレイアウト（1日 = 3列: 開始時刻 / 区切り("-")・備考 / 終了時刻）:
  行1-2   : 日付・曜日（16ブロック、B〜AW列）
  行3-5   : メモ1（日付ごと縦3行マージ）← 日付メモ(schedule_notes)を流し込み
  行6     : メモ2（空欄のまま）
  行7-8   : 朝食見込・夜予約 ← 予約客数(reservation_counts)を流し込み
  行9-10  : キッチン社員 / 行11-28: キッチンA・P
  行29-32 : 2ブロック目ヘッダー（日付・曜日・朝食見込・夜予約=1ブロック目参照式）
  行33-35 : ホール社員 / 行36-51: ホールA・P
  行52-53 : フッター（日付・曜日）
  AX列    : 右側氏名 / AY列: 勤務時間合計式 / AZ列: 人件費式(時給×時間)

従業員数が枠数を超える場合はグループ末尾に行を挿入する（行高・マージも追従）。
日付ブロックは期間日数(月によって13〜16日)に合わせて余りを列削除し、空白を作らない。
"""
from __future__ import annotations
from copy import copy
from datetime import date
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from models.employee import Employee
from models.schedule import SchedulePeriod
from utils.constants import TimeSlot, EmploymentType, PrimaryPosition
from utils.holidays import holiday_set

TEMPLATE_PATH = Path(__file__).parent / "templates" / "SDU_shift_template.xlsx"

DAY_JP = ["月", "火", "水", "木", "金", "土", "日"]

# ── テンプレートのレイアウト定数 ─────────────────────────────────────────
N_BLOCKS     = 16   # テンプレートの日付ブロック数（B〜AW列）。期間日数に合わせて余りは列削除する
COLS_PER_DAY = 3    # 1日 = 3列（開始 / 区切り・備考 / 終了）
C_DATA       = 2    # データ先頭列（B）
C_NAME_L     = 1    # 氏名列（左、A）
NAME_W       = 17.63  # 氏名列の幅
DAY_W        = 5.75   # 日付1列の幅

R_DATE1, R_DOW1   = 1, 2    # 1ブロック目: 日付・曜日
R_MEMO1, R_MEMO2  = 3, 6    # メモ1（3-5行マージ）・メモ2
R_BF, R_DIN       = 7, 8    # 朝食見込・夜予約
K_FT_START, K_FT_SLOTS = 9, 2     # キッチン社員
K_PT_START, K_PT_SLOTS = 11, 18   # キッチンA・P
R_DATE2 = 29                       # 2ブロック目ヘッダー（日付/曜日/朝食見込/夜予約）
H_FT_START, H_FT_SLOTS = 33, 3    # ホール社員
H_PT_START, H_PT_SLOTS = 36, 16   # ホールA・P
R_DATE3 = 52                       # フッター（日付/曜日）

# ── ステータス別のセル装飾（完成イメージ SDU_Shift_full.xlsx に準拠）────
FILL_NONE  = PatternFill(fill_type=None)
FILL_NOTE  = PatternFill("solid", fgColor="FFFF00")  # 黄: 備考メモ
FILL_LEAVE = PatternFill("solid", fgColor="B3E5A1")  # 緑: 有給（3セル全体）
FILL_FTOFF = PatternFill("solid", fgColor="FF0000")  # 赤: 社員の休み・指定休（3セル全体）
FILL_SAT   = PatternFill("solid", fgColor="60CBF3")  # 水色: 土曜の日付・曜日
FILL_SUN   = PatternFill("solid", fgColor="F6C6AC")  # オレンジ: 日曜・祝日の日付・曜日

# 全セル縦横中央揃え（テンプレートの氏名列は標準=左寄せのため明示的に上書き）
ALIGN_CC = Alignment(horizontal="center", vertical="center")


# ── ユーティリティ ────────────────────────────────────────────────────────

def _to_decimal(time_str: str) -> str:
    """'HH:MM' → 小数時刻文字列 ('5.75', '22.5', '13' など)"""
    if not time_str:
        return ""
    try:
        h, m = map(int, time_str.split(":"))
        if m == 0:
            return str(h)
        val = h + m / 60.0
        return f"{val:.2f}".rstrip("0").rstrip(".")
    except Exception:
        return time_str


def _get_shift_cells(
    emp_id: int,
    date_str: str,
    assignments: dict,
    req_map: dict,
    cell_notes: dict,
) -> tuple[str, str, str, str]:
    """
    (開始時刻セル, 区切り/備考セル, 終了時刻セル, スタイル種別) を返す。
    スタイル種別: 'assigned' | 'assigned_note' | 'leave' | 'off' | 'ft_off' | 'am_only' | 'pm_only'

    表示フォーマット:
      通常シフト   : ("13", "-", "22.5")
      備考付きシフト: ("6.5", "オムレツ", "15")
      有給         : ("", "有給", "")
      休み         : ("", "-", "")

    assignments の値は (position_value, is_reinforcement, reinf_start, reinf_end) の4タプル。
    """
    from utils.shift_patterns import PATTERN_MAP

    req  = req_map.get((emp_id, date_str))
    note = (req.note or "").strip() if req else ""
    # スタッフ個別コメント（セル単位のメモ）があれば優先して区切りセルに表示
    cnote = (cell_notes.get((emp_id, date_str)) or "").strip()
    if cnote:
        note = cnote

    # 正社員希望休
    if req and req.pattern_id == "off_request":
        return "", "-", "", "ft_off"

    # 有給チェック（paid_leave パターン or メモに「有給」）
    if (req and req.pattern_id == "paid_leave") or "有給" in note:
        return "", "有給", "", "leave"

    # アサイン確認
    b_raw = assignments.get((emp_id, date_str, TimeSlot.BREAKFAST.value))
    d_raw = assignments.get((emp_id, date_str, TimeSlot.DINNER.value))

    if not b_raw and not d_raw:
        if req and req.pattern_id == "am_only":
            return "", "朝のみ可", "", "am_only"
        if req and req.pattern_id == "pm_only":
            return "", "晩のみ可", "", "pm_only"
        if cnote:
            return "", cnote, "", "assigned_note"
        return "", "-", "", "off"

    # 4タプルから位置情報を展開
    b_pos, b_is_reinf, b_rs, b_re = b_raw if b_raw else (None, False, None, None)
    d_pos, d_is_reinf, d_rs, d_re = d_raw if d_raw else (None, False, None, None)

    # 応援要員は reinf_start/reinf_end を優先して使用
    if b_is_reinf or d_is_reinf:
        if b_raw and not d_raw and b_rs and b_re:
            s, e = _to_decimal(b_rs), _to_decimal(b_re)
        elif d_raw and not b_raw and d_rs and d_re:
            s, e = _to_decimal(d_rs), _to_decimal(d_re)
        elif b_raw and d_raw:
            s = _to_decimal(b_rs) if b_rs else "6"
            e = _to_decimal(d_re) if d_re else "23"
        else:
            s, e = _slot_default(b_pos, d_pos)
        if note:
            return s, note, e, "assigned_note"
        return s, "-", e, "assigned"

    # 時刻取得（パターンから）
    if req and req.pattern_id == "double":
        s, e = "6", "23"
    elif req and req.pattern_id == "custom" and req.custom_start and req.custom_end:
        s = _to_decimal(req.custom_start)
        e = _to_decimal(req.custom_end)
    elif req and req.pattern_id:
        p = PATTERN_MAP.get(req.pattern_id)
        if p and p.start and p.end:
            s = _to_decimal(p.start)
            e = _to_decimal(p.end)
        else:
            s, e = _slot_default(b_pos, d_pos)
    else:
        s, e = _slot_default(b_pos, d_pos)

    # 備考を時刻の間（区切りセル）に挟む形式
    if note:
        return s, note, e, "assigned_note"
    return s, "-", e, "assigned"


def _slot_default(b_pos, d_pos) -> tuple[str, str]:
    """ポジション別のデフォルト勤務時間を返す（HH:MM を小数時刻に変換）"""
    from db import repositories as repo
    pos_key = b_pos or d_pos or "hall"  # ポジション名 ("hall" or "kitchen")
    if b_pos and d_pos:
        s = _to_decimal(repo.get_app_setting(f"ft_{pos_key}_breakfast_start", "06:00"))
        e = _to_decimal(repo.get_app_setting(f"ft_{pos_key}_dinner_end",      "22:00"))
    elif b_pos:
        s = _to_decimal(repo.get_app_setting(f"ft_{pos_key}_breakfast_start", "06:00"))
        e = _to_decimal(repo.get_app_setting(f"ft_{pos_key}_breakfast_end",   "11:00"))
    else:
        s = _to_decimal(repo.get_app_setting(f"ft_{pos_key}_dinner_start", "17:00"))
        e = _to_decimal(repo.get_app_setting(f"ft_{pos_key}_dinner_end",   "22:00"))
    return s, e


def _num_or_text(v: str):
    """数値に変換できれば float/int、できなければ文字列のまま返す"""
    if v == "":
        return None
    try:
        f = float(v)
        return int(f) if f.is_integer() else f
    except ValueError:
        return v


def _insert_rows_keep_layout(ws, idx: int, amount: int, max_col: int):
    """idx 行の直前に amount 行挿入する。
    openpyxl の insert_rows はマージセル・行高を追従させないため手動でずらし、
    挿入行のスタイル・行高は直前の行（同グループの枠）からコピーする。"""
    if amount <= 0:
        return
    # マージ範囲の退避とシフト
    to_shift = [r for r in list(ws.merged_cells.ranges) if r.min_row >= idx]
    for r in to_shift:
        ws.unmerge_cells(str(r))
    ws.insert_rows(idx, amount)
    for r in to_shift:
        ws.merge_cells(start_row=r.min_row + amount, end_row=r.max_row + amount,
                       start_column=r.min_col, end_column=r.max_col)
    # 行高のシフト（下から処理して上書きを防ぐ）
    heights = {r: dim.height for r, dim in ws.row_dimensions.items()
               if r >= idx and dim.height is not None}
    for r in sorted(heights, reverse=True):
        ws.row_dimensions[r + amount].height = heights[r]
    # 挿入行: 直前の行からスタイル・行高をコピー
    donor = idx - 1
    donor_h = ws.row_dimensions[donor].height
    for r in range(idx, idx + amount):
        if donor_h:
            ws.row_dimensions[r].height = donor_h
        for c in range(1, max_col + 1):
            ws.cell(r, c)._style = copy(ws.cell(donor, c)._style)


def _delete_trailing_blocks(ws, n: int):
    """期間日数 n に合わせて余った日付ブロック列を削除する（空白ブロックを作らない）。
    openpyxl の delete_cols はマージセルを追従させないため手動でずらす。"""
    extra = N_BLOCKS - n
    if extra <= 0:
        return
    first  = C_DATA + n * COLS_PER_DAY   # 削除開始列
    amount = extra * COLS_PER_DAY
    for rng in list(ws.merged_cells.ranges):
        if rng.min_col >= first + amount:
            # 削除範囲より右（氏名列など）→ 左へシフト
            ws.unmerge_cells(str(rng))
            ws.merge_cells(start_row=rng.min_row, end_row=rng.max_row,
                           start_column=rng.min_col - amount,
                           end_column=rng.max_col - amount)
        elif rng.min_col >= first:
            # 削除範囲内のブロックマージ → 解除
            ws.unmerge_cells(str(rng))
    ws.delete_cols(first, amount)
    # 列幅を再構築（テンプレートのグループ幅指定は削除に追従しないため）
    for letter in list(ws.column_dimensions.keys()):
        del ws.column_dimensions[letter]
    c_name_r = C_DATA + n * COLS_PER_DAY
    ws.column_dimensions[get_column_letter(C_NAME_L)].width = NAME_W
    for c in range(C_DATA, c_name_r):
        ws.column_dimensions[get_column_letter(c)].width = DAY_W
    ws.column_dimensions[get_column_letter(c_name_r)].width = NAME_W


def _block_col(i: int) -> int:
    """日付ブロック i (0始まり) の先頭列"""
    return C_DATA + i * COLS_PER_DAY


def _hours_formula(row: int, n: int) -> str:
    """=(D9-B9)+(G9-E9)+... 形式の勤務時間合計式（期間 n 日分）を返す"""
    terms = []
    for i in range(n):
        c = _block_col(i)
        terms.append(f"({get_column_letter(c + 2)}{row}-{get_column_letter(c)}{row})")
    return "=" + "+".join(terms)


# ── メイン出力 ────────────────────────────────────────────────────────────

def export_excel(
    path: str,
    period: SchedulePeriod,
    employees: list[Employee],
    assignments: dict,
):
    """SDU_shift テンプレートにシフトデータを流し込んで Excel を出力する"""
    from db import repositories as repo
    requests = repo.get_shift_requests(period.id)
    req_map  = {(r.employee_id, r.date): r for r in requests}
    notes    = repo.get_schedule_notes(period.id)        # 日付メモ → メモ1欄
    reserves = repo.get_reservation_counts(period.id)    # 予約客数 → 朝食見込・夜予約
    cell_notes = repo.get_cell_notes(period.id)          # スタッフ個別コメント → 区切りセル

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    dates    = list(period.date_range())[:N_BLOCKS]  # テンプレートは最大16日分
    n        = len(dates)
    holidays = holiday_set(dates)

    # 月の日数(28/30/31日)に合わせて余りブロックを列ごと削除し、空白を作らない
    _delete_trailing_blocks(ws, n)
    c_name_r = C_DATA + n * COLS_PER_DAY  # 氏名列（右）
    c_hours  = c_name_r + 1               # 勤務時間合計式
    c_wage   = c_name_r + 2               # 人件費式
    # 削除で左に寄ってきたテンプレート外の作業メモ等を消す
    for row in ws.iter_rows(min_row=1, max_row=100, min_col=c_wage + 1, max_col=63):
        for cell in row:
            if cell.value is not None:
                cell.value = None

    # ── 従業員をテンプレートの4グループに分類 ─────────────────────────
    # output_position 優先、未設定なら primary_position で判断
    def _out_pos(e):
        return e.output_position or e.primary_position

    def _is_ft(e):
        return e.employment_type == EmploymentType.FULL_TIME

    kitchen = [e for e in employees if _out_pos(e) == PrimaryPosition.KITCHEN]
    hall    = [e for e in employees if _out_pos(e) != PrimaryPosition.KITCHEN]
    k_ft = [e for e in kitchen if _is_ft(e)]
    k_pt = [e for e in kitchen if not _is_ft(e)]
    h_ft = [e for e in hall if _is_ft(e)]
    h_pt = [e for e in hall if not _is_ft(e)]

    # ── 枠が足りないグループには行を挿入（下のグループから処理）────────
    k_ft_extra = max(0, len(k_ft) - K_FT_SLOTS)
    k_pt_extra = max(0, len(k_pt) - K_PT_SLOTS)
    h_ft_extra = max(0, len(h_ft) - H_FT_SLOTS)
    h_pt_extra = max(0, len(h_pt) - H_PT_SLOTS)
    _insert_rows_keep_layout(ws, H_PT_START + H_PT_SLOTS, h_pt_extra, c_wage)
    _insert_rows_keep_layout(ws, H_FT_START + H_FT_SLOTS, h_ft_extra, c_wage)
    _insert_rows_keep_layout(ws, K_PT_START + K_PT_SLOTS, k_pt_extra, c_wage)
    _insert_rows_keep_layout(ws, K_FT_START + K_FT_SLOTS, k_ft_extra, c_wage)

    # 挿入後の行位置
    ko       = k_ft_extra + k_pt_extra
    k_pt_row = K_PT_START + k_ft_extra
    r_date2  = R_DATE2 + ko
    h_ft_row = H_FT_START + ko
    h_pt_row = H_PT_START + ko + h_ft_extra
    r_date3  = R_DATE3 + ko + h_ft_extra + h_pt_extra

    # ── タイトル・シート名 ─────────────────────────────────────────────
    start_d = date.fromisoformat(period.start_date)
    end_d   = date.fromisoformat(period.end_date)
    half    = "前半" if start_d.day <= 15 else "後半"
    title   = f"{start_d.month}月      　　　　 {half}"
    for r in (R_DATE1, r_date2, r_date3):
        for c in (C_NAME_L, c_name_r):
            cell = ws.cell(r, c)
            cell.value = title
            cell.alignment = ALIGN_CC
    ws.title = (f"{start_d.month:02d}{start_d.day:02d}_{end_d.month:02d}{end_d.day:02d}"
                + ("(確定)" if period.status == "confirmed" else ""))

    # ── ヘッダー・フッター（日付/曜日/メモ/朝食見込/夜予約）────────────
    for i, d in enumerate(dates):
        col = _block_col(i)
        cl  = get_column_letter(col)
        ds = d.isoformat()
        day_val = d.day
        dow     = DAY_JP[d.weekday()]
        note    = notes.get(ds, "") or None
        res     = reserves.get(ds, {})
        bf      = res.get("breakfast")
        din     = res.get("dinner")
        if d.weekday() == 5:
            day_fill = FILL_SAT
        elif d.weekday() == 6 or ds in holidays:
            day_fill = FILL_SUN
        else:
            day_fill = FILL_NONE

        for r in (R_DATE1, r_date2, r_date3):
            ws.cell(r, col).value = day_val
            ws.cell(r + 1, col).value = dow
            # 土曜=水色 / 日曜・祝日=オレンジ（マージ範囲の3セルすべてに塗る）
            for cc in range(col, col + COLS_PER_DAY):
                ws.cell(r, cc).fill = day_fill
                ws.cell(r + 1, cc).fill = day_fill
        ws.cell(R_MEMO1, col).value = note
        ws.cell(R_MEMO2, col).value = None
        ws.cell(R_BF,  col).value = bf
        ws.cell(R_DIN, col).value = din
        # 2ブロック目の朝食見込・夜予約は1ブロック目を参照する式（テンプレート流儀）
        ws.cell(r_date2 + 2, col).value = f"={cl}{R_BF}"  if bf  is not None else None
        ws.cell(r_date2 + 3, col).value = f"={cl}{R_DIN}" if din is not None else None

    # 2ブロック目の朝食見込・夜予約ラベルとヘッダーラベルの中央揃え
    for r, label in ((R_BF, "朝食見込"), (R_DIN, "夜予約"),
                     (r_date2 + 2, "朝食見込"), (r_date2 + 3, "夜予約")):
        for c in (C_NAME_L, c_name_r):
            cell = ws.cell(r, c)
            cell.value = label
            cell.alignment = ALIGN_CC

    # ── 従業員グループの流し込み ───────────────────────────────────────
    def _fill_group(start_row: int, slot_count: int, emps: list[Employee]):
        for j in range(slot_count):
            r   = start_row + j
            emp = emps[j] if j < len(emps) else None
            for c in (C_NAME_L, c_name_r):
                cell = ws.cell(r, c)
                cell.value = emp.name if emp else None
                cell.alignment = ALIGN_CC
            # 全ブロックをクリア（テンプレートの凡例サンプル・前回値を消す）
            for i in range(n):
                col = _block_col(i)
                for cc in range(col, col + COLS_PER_DAY):
                    cell = ws.cell(r, cc)
                    cell.value = None
                    cell.fill  = FILL_NONE
            if emp is None:
                ws.cell(r, c_hours).value = None
                ws.cell(r, c_wage).value  = None
                continue
            for i, d in enumerate(dates):
                col = _block_col(i)
                s_txt, m_txt, e_txt, style = _get_shift_cells(
                    emp.id, d.isoformat(), assignments, req_map, cell_notes)
                ws.cell(r, col).value     = _num_or_text(s_txt)
                ws.cell(r, col + 1).value = m_txt or None
                ws.cell(r, col + 2).value = _num_or_text(e_txt)
                if style in ("assigned_note", "am_only", "pm_only"):
                    ws.cell(r, col + 1).fill = FILL_NOTE
                elif style == "leave":
                    # 有給は3セル全体を緑塗り（完成イメージ準拠）
                    for cc in range(col, col + COLS_PER_DAY):
                        ws.cell(r, cc).fill = FILL_LEAVE
                elif style == "ft_off" or (
                    style == "off" and emp.employment_type == EmploymentType.FULL_TIME
                ):
                    # 社員の休み・指定休は3セル全体を赤塗り
                    for cc in range(col, col + COLS_PER_DAY):
                        ws.cell(r, cc).fill = FILL_FTOFF
            ws.cell(r, c_hours).value = _hours_formula(r, n)
            ws.cell(r, c_wage).value = (
                f"={get_column_letter(c_hours)}{r}*{emp.hourly_wage}"
                if emp.hourly_wage else None)

    _fill_group(K_FT_START, K_FT_SLOTS + k_ft_extra, k_ft)
    _fill_group(k_pt_row,   K_PT_SLOTS + k_pt_extra, k_pt)
    _fill_group(h_ft_row,   H_FT_SLOTS + h_ft_extra, h_ft)
    _fill_group(h_pt_row,   H_PT_SLOTS + h_pt_extra, h_pt)

    wb.save(path)
